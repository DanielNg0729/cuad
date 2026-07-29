"""
Flask server for the CUAD contract highlighter UI.

Two-phase flow so the loading screen can show real stages and so switching the
model doesn't re-parse the PDF:

    POST /api/parse    multipart {file}        -> {doc_id, text, chars}
    POST /api/extract  json {doc_id, model}    -> {entities, spans_by_label, ...}

Other routes:
    GET  /                 main page
    GET  /categories       reference page (41 categories)
    GET  /api/models       model dropdown contents
    GET  /api/categories   41 categories with colours/descriptions

Run:
    cd webui && python app.py        # http://127.0.0.1:5000
"""

import io
import threading
import uuid

from dotenv import load_dotenv
from flask import Flask, jsonify, request, send_file, send_from_directory

import extract
import rag
import verify

# .env lives at the project root; load it so OPENAI_API_KEY / GROQ_API_KEY exist.
load_dotenv(extract.PROJECT_ROOT / ".env")

app = Flask(__name__, static_folder="static", static_url_path="/static")


@app.after_request
def _no_cache(resp):
    """Never let the browser cache the HTML/JS/CSS. Otherwise an old cached
    app.js keeps talking to an updated backend and silently breaks (e.g. after
    the /api/extract response shape changed for the progress bar)."""
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp

# In-memory store of parsed docs: doc_id -> {"text": str, "chunks": [str]}.
# Fine for a local single-user tool; restart clears it.
_DOCS: dict[str, dict] = {}
MAX_DOCS = 20

# Extraction jobs run in a background thread so the UI can poll live progress.
# job_id -> {stage, completed, total, ok, failed, finished, result, error}
_JOBS: dict[str, dict] = {}
MAX_JOBS = 40

# Human review state per document, so Approve/Reject/Edit decisions and
# manually-added answers survive page reloads (until the server restarts).
#   doc_id -> {item_id -> {id, label, text, original, status, page}}
# status: pending | approved | rejected | edited | manual.
# `original` is the verbatim LLM span (None for manual answers); merge on
# re-extraction is keyed by (label, original) so decisions aren't lost or
# duplicated when the same span is found again.
_REVIEWS: dict[str, dict[str, dict]] = {}
_VALID_STATUS = {"pending", "approved", "rejected", "edited", "manual"}


# --- pages ------------------------------------------------------------------

@app.get("/")
def index():
    return send_from_directory("static", "index.html")


@app.get("/categories")
def categories_page():
    return send_from_directory("static", "categories.html")


# --- metadata APIs ----------------------------------------------------------

@app.get("/api/models")
def api_models():
    return jsonify([{"id": m["id"], "label": m["label"], "provider": m["provider"]}
                    for m in extract.MODELS])


@app.get("/api/categories")
def api_categories():
    return jsonify(extract.load_categories())


@app.post("/api/categories")
def api_add_category():
    """Add a user-defined category. Body: {label, description, answer_format?}.
    The description is used as the model prompt, so the next extraction (Analyze
    / Re-run model) will find and highlight it just like the 41 built-ins."""
    data = request.get_json(silent=True) or {}
    try:
        cat = extract.add_custom_category(
            data.get("label"), data.get("description"), data.get("answer_format", "")
        )
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify(cat), 201


@app.delete("/api/categories/<path:label>")
def api_delete_category(label):
    """Remove a custom category by label. Built-in categories can't be removed."""
    if extract.remove_custom_category(label):
        return jsonify({"ok": True})
    return jsonify({"error": "Unknown or non-custom category."}), 404


# --- pipeline APIs ----------------------------------------------------------

@app.post("/api/parse")
def api_parse():
    """Phase 1: PDF -> docling markdown -> chunks. Returns the text to render and
    a doc_id to reuse for extraction."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded (field name must be 'file')."}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "Empty filename."}), 400

    pdf_bytes = f.read()
    if not pdf_bytes:
        return jsonify({"error": "Uploaded file is empty."}), 400

    try:
        text, chunks = extract.parse_pdf_to_text(pdf_bytes, f.filename)
    except Exception as e:  # noqa: BLE001
        return jsonify({"error": f"PDF parsing failed: {e}"}), 500

    if not text.strip():
        return jsonify({"error": "No text could be extracted from this PDF."}), 422

    doc_id = uuid.uuid4().hex
    # bound memory: drop the oldest doc (and its review state) if over the cap
    if len(_DOCS) >= MAX_DOCS:
        oldest = next(iter(_DOCS))
        _DOCS.pop(oldest)
        _REVIEWS.pop(oldest, None)
    # keep the raw PDF bytes so the front-end can render the actual document
    _DOCS[doc_id] = {"text": text, "chunks": chunks, "filename": f.filename,
                     "pdf": pdf_bytes}

    return jsonify({
        "doc_id": doc_id,
        "filename": f.filename,
        "chars": len(text),
        "n_chunks": len(chunks),
    })


@app.get("/api/pdf/<doc_id>")
def api_pdf(doc_id):
    """Serve the original uploaded PDF so PDF.js can render it in the browser."""
    from flask import Response
    doc = _DOCS.get(doc_id)
    if doc is None or "pdf" not in doc:
        return jsonify({"error": "Unknown or expired doc_id."}), 404
    return Response(doc["pdf"], mimetype="application/pdf")


def _run_extract_job(job_id: str, doc: dict, model_id: str, mode: str):
    """Background worker: runs extraction (scan or RAG), streaming progress into
    the job and storing the final result (or error) when done."""
    job = _JOBS[job_id]

    def progress(completed, total, ok, failed):
        job.update(completed=completed, total=total, ok=ok, failed=failed)

    try:
        if mode in ("rag", "hybrid"):
            # Embed the chunks once and cache the matrix on the doc, so switching
            # the LLM or re-running RAG doesn't pay to re-embed the same chunks.
            # (Hybrid mode still needs these: it cosine-reranks its BM25 shortlist.)
            emb = doc.get("chunk_emb")
            if emb is None:
                job.update(stage="embedding", completed=0, total=len(doc["chunks"]),
                           ok=0, failed=0)
                emb, _ = rag.embed_texts(
                    doc["chunks"],
                    progress_cb=lambda done, tot: job.update(completed=done, total=tot),
                )
                doc["chunk_emb"] = emb
            # switch to the per-category extraction phase (its own progress scale)
            job.update(stage="extracting", completed=0,
                       total=len(extract.load_categories()))
            if mode == "hybrid":
                result = rag.rag_extract(doc["text"], doc["chunks"], model_id,
                                         chunk_embeddings=emb, progress_cb=progress,
                                         search="hybrid", top_k=rag.HYBRID_TOP_K,
                                         hybrid_n=rag.HYBRID_BM25_N)
            else:
                result = rag.rag_extract(doc["text"], doc["chunks"], model_id,
                                         chunk_embeddings=emb, progress_cb=progress)
            unit, total = "categor(y/ies)", result.get("n_labels", 0)
            # Remember how this label's chunks were retrieved, so a later "AI
            # verify" pass for the same label can silently reuse the exact same
            # top-k retrieval instead of a generic +/- char window (see api_verify).
            doc["retrieval_meta"] = {
                "search": result.get("search", "cosine"),
                "top_k": result.get("top_k", rag.DEFAULT_TOP_K),
                "hybrid_n": result.get("hybrid_n"),
            }
        else:
            result = extract.extract(doc["text"], doc["chunks"], model_id, progress_cb=progress)
            unit, total = "chunk(s)", result.get("n_chunks", 0)
            doc["retrieval_meta"] = None   # full scan: no per-label retrieval to mirror

        if result.get("n_ok", 0) == 0 and result.get("n_failed", 0) > 0:
            # everything failed -> the model never really ran (dead id, bad key)
            job["error"] = (
                f"Model '{model_id}' failed on all {result['n_failed']} {unit}. "
                f"{result.get('error') or ''}".strip()
            )
        else:
            if result.get("n_failed", 0) > 0:
                result["warning"] = (
                    f"{result['n_failed']} of {total} {unit} failed; "
                    f"results may be incomplete."
                )
            job["result"] = result
    except Exception as e:  # noqa: BLE001
        job["error"] = f"Extraction failed: {e}"
    finally:
        job["finished"] = True


@app.post("/api/extract")
def api_extract():
    """Phase 2: start a background extraction job and return its id immediately.
    The client polls /api/extract_status/<job_id> for live progress."""
    data = request.get_json(silent=True) or {}
    doc_id = data.get("doc_id")
    model_id = data.get("model")
    mode = data.get("mode", "scan")   # "scan" (all chunks), "rag" (cosine top-k), or
                                       # "hybrid" (BM25 top-5 -> cosine top-5)

    doc = _DOCS.get(doc_id)
    if doc is None:
        return jsonify({"error": "Unknown or expired doc_id. Re-upload the PDF."}), 404
    if model_id not in {m["id"] for m in extract.MODELS}:
        return jsonify({"error": f"Unknown model: {model_id!r}"}), 400
    if mode not in {"scan", "rag", "hybrid"}:
        return jsonify({"error": f"Unknown mode: {mode!r}"}), 400

    # progress denominator: chunks for a full scan, categories for RAG/hybrid.
    total = len(doc["chunks"]) if mode == "scan" else len(extract.load_categories())

    job_id = uuid.uuid4().hex
    if len(_JOBS) >= MAX_JOBS:
        _JOBS.pop(next(iter(_JOBS)))
    _JOBS[job_id] = {
        "stage": "extracting", "completed": 0, "total": total,
        "ok": 0, "failed": 0, "finished": False, "result": None, "error": None,
        "model": model_id, "mode": mode,
    }
    threading.Thread(
        target=_run_extract_job, args=(job_id, doc, model_id, mode), daemon=True
    ).start()
    return jsonify({"job_id": job_id, "total": total, "mode": mode})


@app.get("/api/extract_status/<job_id>")
def api_extract_status(job_id):
    """Poll target: returns live counts; embeds the final result once finished."""
    job = _JOBS.get(job_id)
    if job is None:
        return jsonify({"error": "Unknown job id."}), 404
    payload = {
        "stage": job["stage"], "completed": job["completed"], "total": job["total"],
        "ok": job["ok"], "failed": job["failed"], "finished": job["finished"],
        "mode": job.get("mode", "scan"),
    }
    if job["finished"]:
        if job["error"]:
            payload["error"] = job["error"]
        else:
            payload["result"] = job["result"]
    return jsonify(payload)


# --- review APIs ------------------------------------------------------------

@app.get("/api/review/<doc_id>")
def api_review_get(doc_id):
    """Current review items for a document (used to restore state after reload)."""
    if doc_id not in _DOCS:
        return jsonify({"error": "Unknown or expired doc_id."}), 404
    return jsonify({"items": list(_REVIEWS.get(doc_id, {}).values())})


@app.post("/api/review/<doc_id>/sync")
def api_review_sync(doc_id):
    """Merge the model's extracted answers into the stored review state.

    Body: {"answers": [{"label", "text", "page"}]}. Existing decisions are kept
    (matched by (label, original)); new spans are added as 'pending'; stale
    'pending' spans from a previous run that this run didn't produce are dropped.
    Manual answers and reviewed (approved/rejected/edited) items are preserved.
    Returns the full item list."""
    if doc_id not in _DOCS:
        return jsonify({"error": "Unknown or expired doc_id."}), 404
    data = request.get_json(silent=True) or {}
    answers = data.get("answers", []) or []

    store = _REVIEWS.setdefault(doc_id, {})
    by_key = {(it["label"], it["original"]): it
              for it in store.values() if it["original"] is not None}

    incoming_keys = set()
    for a in answers:
        label = a.get("label")
        text = (a.get("text") or "").strip()
        if not label or not text:
            continue
        key = (label, text)
        incoming_keys.add(key)
        if key in by_key:
            if a.get("page") is not None:
                by_key[key]["page"] = a.get("page")   # refresh page mapping
            continue
        iid = uuid.uuid4().hex[:12]
        store[iid] = {"id": iid, "label": label, "text": text,
                      "original": text, "status": "pending", "page": a.get("page")}

    # drop stale 'pending' model spans this run no longer found
    for iid in list(store):
        it = store[iid]
        if (it["original"] is not None and it["status"] == "pending"
                and (it["label"], it["original"]) not in incoming_keys):
            del store[iid]

    return jsonify({"items": list(store.values())})


@app.patch("/api/review/<doc_id>/<item_id>")
def api_review_update(doc_id, item_id):
    """Update one item: status (approve/reject), text (edit), and/or page.
    Editing the text moves the item to 'edited' (or stays 'manual')."""
    store = _REVIEWS.get(doc_id)
    if not store or item_id not in store:
        return jsonify({"error": "Unknown doc_id or item_id."}), 404
    it = store[item_id]
    data = request.get_json(silent=True) or {}

    if "text" in data:
        text = (data.get("text") or "").strip()
        if not text:
            return jsonify({"error": "Edited text cannot be empty."}), 400
        it["text"] = text
        it["status"] = "manual" if it["original"] is None else "edited"
        it.pop("verify", None)   # editing the text invalidates any prior AI verdict
    if "status" in data:
        st = data["status"]
        if st not in _VALID_STATUS:
            return jsonify({"error": f"Invalid status: {st!r}"}), 400
        it["status"] = st
    if "page" in data:
        it["page"] = data["page"]
    return jsonify(it)


@app.post("/api/review/<doc_id>")
def api_review_add(doc_id):
    """Add a manual answer for a label. Body: {"label", "text", "page"}."""
    if doc_id not in _DOCS:
        return jsonify({"error": "Unknown or expired doc_id."}), 404
    data = request.get_json(silent=True) or {}
    label = data.get("label")
    text = (data.get("text") or "").strip()
    if not label or not text:
        return jsonify({"error": "label and text are required."}), 400
    store = _REVIEWS.setdefault(doc_id, {})
    iid = uuid.uuid4().hex[:12]
    store[iid] = {"id": iid, "label": label, "text": text,
                  "original": None, "status": "manual", "page": data.get("page")}
    return jsonify(store[iid])


@app.delete("/api/review/<doc_id>/<item_id>")
def api_review_delete(doc_id, item_id):
    """Remove an item (used to discard a manually-added answer)."""
    store = _REVIEWS.get(doc_id)
    if not store or item_id not in store:
        return jsonify({"error": "Unknown doc_id or item_id."}), 404
    del store[item_id]
    return jsonify({"ok": True})


# --- AI verification --------------------------------------------------------

@app.post("/api/verify/<doc_id>")
def api_verify(doc_id):
    """Second-opinion pass for one category. Body: {"label", "model"}.

    Asks the chosen model whether each (non-rejected) answer under `label` is a
    correct instance of that category, shown with its surrounding document text.
    Stores the verdict on each review item (survives reload) and returns the
    updated items so the UI can render a badge without a full refetch."""
    doc = _DOCS.get(doc_id)
    if doc is None:
        return jsonify({"error": "Unknown or expired doc_id."}), 404
    data = request.get_json(silent=True) or {}
    label = data.get("label")
    model_id = data.get("model")
    if not label:
        return jsonify({"error": "label is required."}), 400
    if model_id not in {m["id"] for m in extract.MODELS}:
        return jsonify({"error": f"Unknown model: {model_id!r}"}), 400

    desc = next((c["description"] for c in extract.load_categories()
                 if c["label"] == label), "")
    store = _REVIEWS.get(doc_id, {})
    items = [it for it in store.values()
             if it["label"] == label and it["status"] != "rejected"]
    if not items:
        return jsonify({"items": [], "message": "No answers to verify for this label."})

    # If the document was last extracted with RAG/hybrid, silently reuse that
    # SAME top-k retrieval for this label instead of a generic +/- char window,
    # so the verifier judges answers against exactly what the extractor saw.
    shared_context = None
    meta = doc.get("retrieval_meta")
    if meta and doc.get("chunk_emb") is not None:
        try:
            shared_context, bm25 = rag.retrieve_context_for_label(
                label, desc, doc["chunks"], doc["chunk_emb"],
                search=meta["search"], top_k=meta["top_k"],
                hybrid_n=meta.get("hybrid_n") or rag.HYBRID_BM25_N,
                bm25=doc.get("bm25"),
            )
            if bm25 is not None:
                doc["bm25"] = bm25   # cache so later verify calls on this doc skip rebuilding it
        except Exception:  # noqa: BLE001
            shared_context = None   # fall back to the per-answer window below

    try:
        verdicts = verify.verify_label(
            label, desc, [{"id": it["id"], "text": it["text"]} for it in items],
            doc["text"], model_id, shared_context=shared_context,
        )
    except Exception as e:  # noqa: BLE001
        return jsonify({"error": f"Verification failed: {e}"}), 500

    updated = []
    for v in verdicts:
        it = store.get(v["id"])
        if it is not None:
            it["verify"] = {"verdict": v["verdict"], "reason": v["reason"], "model": model_id}
            updated.append(it)
    return jsonify({"items": updated})


# --- export -----------------------------------------------------------------

_EXPORT_STATUS_LABEL = {"pending": "unreviewed"}   # rename for the spreadsheet


def _build_export_workbook(rows: list[dict], filename: str | None):
    """An .xlsx with one row per kept answer: Category | Answer | Status | Page."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Highlights"

    headers = ["Category", "Answer", "Status", "AI Verdict", "Page"]
    head_fill = PatternFill("solid", fgColor="3B6FD4")
    head_font = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center")

    for r, it in enumerate(rows, start=2):
        status = _EXPORT_STATUS_LABEL.get(it["status"], it["status"])
        v = it.get("verify") or {}
        verdict = f"{v['verdict']} — {v['reason']}" if v.get("verdict") else ""
        ws.cell(row=r, column=1, value=it["label"])
        ans = ws.cell(row=r, column=2, value=it["text"])
        ans.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=3, value=status)
        vcell = ws.cell(row=r, column=4, value=verdict)
        vcell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=5, value=it.get("page"))

    widths = [34, 70, 12, 40, 7]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return wb


@app.get("/api/export/<doc_id>")
def api_export(doc_id):
    """Download the reviewed highlights as Excel. Rejected answers are excluded;
    rows are ordered by category (CSV order), then answer text."""
    doc = _DOCS.get(doc_id)
    if doc is None:
        return jsonify({"error": "Unknown or expired doc_id."}), 404

    store = _REVIEWS.get(doc_id, {})
    order = {c["label"]: i for i, c in enumerate(extract.load_categories())}
    rows = [it for it in store.values() if it["status"] != "rejected"]
    rows.sort(key=lambda it: (order.get(it["label"], 1_000), it["text"].lower()))

    wb = _build_export_workbook(rows, doc.get("filename"))
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    base = (doc.get("filename") or "highlights").rsplit(".", 1)[0]
    download_name = f"{base}_highlights.xlsx"
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=download_name,
    )


if __name__ == "__main__":
    # threaded so the long parse/extract request doesn't block static assets.
    app.run(host="127.0.0.1", port=5000, debug=True, threaded=True)
