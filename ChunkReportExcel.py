"""Build an Excel report of how many chunks each contract was split into.

Reads test_chunking.json (produced by chunking.py) and writes
chunk_report.xlsx with:
  * one row per contract (number, contract_id, num_chunks)
  * a totals/summary block
  * a bar chart visualising num_chunks across all 102 contracts.

Usage:
    python ChunkReportExcel.py
    python ChunkReportExcel.py --in test_chunking.json --out chunk_report.xlsx
"""

import argparse
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--in", dest="infile", default="test_chunking.json",
                    help="Chunk file to report on (default test_chunking.json).")
    ap.add_argument("--out", dest="outfile", default="chunk_report.xlsx",
                    help="Excel file to write (default chunk_report.xlsx).")
    args = ap.parse_args()

    doc = json.loads(Path(args.infile).read_text(encoding="utf-8"))
    contracts = doc["data"]

    def word_count(chunks):
        # Count words across all of a contract's chunk text. \w+ tokens only, so
        # Markdown markers like ## and ** don't get counted as words.
        return len(re.findall(r"\w+", " ".join(chunks)))

    # num_chunks is stored per contract; fall back to len(chunks) just in case.
    rows = [(c["contract_id"],
             c.get("num_chunks", len(c["chunks"])),
             word_count(c["chunks"]))
            for c in contracts]

    wb = Workbook()
    ws = wb.active
    ws.title = "Chunks per contract"

    # --- header row -------------------------------------------------------
    headers = ["#", "Contract", "Num chunks", "Word count"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="4472C4")
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")

    # --- data rows --------------------------------------------------------
    for i, (cid, n, w) in enumerate(rows, 1):
        ws.append([i, cid, n, w])

    first_data, last_data = 2, len(rows) + 1

    # --- summary block (to the side of the table) -------------------------
    nums = [n for _, n, _ in rows]
    words = [w for _, _, w in rows]
    summary = [
        ("Contracts", len(rows)),
        ("Total chunks", sum(nums)),
        ("Avg chunks / contract", round(sum(nums) / len(rows), 1)),
        ("Min / Max chunks", f"{min(nums)} / {max(nums)}"),
        ("Total words", sum(words)),
        ("Avg words / contract", round(sum(words) / len(rows))),
        ("Min / Max words", f"{min(words)} / {max(words)}"),
    ]
    for r, (label, val) in enumerate(summary, start=1):
        ws.cell(row=r, column=6, value=label).font = Font(bold=True)
        ws.cell(row=r, column=7, value=val)

    # --- column widths & number alignment ---------------------------------
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 14
    for col in (1, 3, 4):
        for row in ws.iter_rows(min_row=first_data, max_row=last_data,
                                min_col=col, max_col=col):
            row[0].alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    # --- bar chart --------------------------------------------------------
    chart = BarChart()
    chart.type = "col"
    chart.title = "Number of chunks per contract"
    chart.y_axis.title = "Chunks"
    chart.x_axis.title = "Contract #"
    chart.height = 10
    chart.width = 34
    chart.legend = None

    data = Reference(ws, min_col=3, min_row=1, max_row=last_data)   # incl. header
    cats = Reference(ws, min_col=1, min_row=first_data, max_row=last_data)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.gapWidth = 30
    # Anchor the chart below the summary, to the right of the table.
    ws.add_chart(chart, "E8")

    wb.save(args.outfile)
    print(f"Wrote {args.outfile}: {len(rows)} contracts, "
          f"{sum(nums)} chunks total (avg {sum(nums)/len(rows):.1f}, "
          f"min {min(nums)}, max {max(nums)}).")


if __name__ == "__main__":
    main()
